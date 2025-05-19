import pandas as pd
import logging
import warnings
import numpy as np
from openpyxl import load_workbook

'''
sheet/ws => 1:学生信息/2：大纲要求/3：学生成绩/4：考试成绩/5：达成度汇总/6：目标明细
'''
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s  %(levelname)s:%(message)s')
adress = r"C:\Users\lenovo\Desktop\课程开发平台试验excel表.xlsx"
# 使用 pd.read_excel() 读取时,默认会将 Excel 的第 1 行作为表头
sheet1 = pd.read_excel(adress, sheet_name='学生成绩')
sheet2 = pd.read_excel(adress, sheet_name='大纲要求')
sheet3 = pd.read_excel(adress, sheet_name='学生成绩')
sheet4 = pd.read_excel(adress, sheet_name='考试成绩')
student_number = sheet1.shape[0]
print('student_number: ', student_number)
# 屏蔽 Pandas 关于 applymap() 方法已被弃用的警告信息
warnings.filterwarnings('ignore', category=FutureWarning)


def index(letter):
    '''将以字母为列索引转换成以数字为列索引'''
    return (len(letter) - 1) * 26 + ord(letter[-1]) - 65


def zs_level_to_score():
    '''
    从大纲要求中读取作业和测试的等级分数对应表，对学生成绩里作业和实验进行等级
    到分数的转换，并存为新的列
    return:作业和实验为分数制的DataFrame
    '''
    # 等级分数对应表存为一个字典dict_level_score
    dict_level_score = {
        sheet2.iloc[row, index('I')]: int(sheet2.iloc[row, index('J')]) for row
        in range(0, 4)}
    level_df = sheet3.iloc[0:student_number,
               index('E'):index('N') + 1]  # level为作业和实验为等级制的DataFrame
    # 一次性映射所有列,将等级映射为分数
    score_df = level_df.applymap(
        lambda x: dict_level_score.get(x, 0) if pd.notna(
            x) else 0)  # score为分数制的DataFrame
    logging.info('successfully run the zs_level_to_score')
    return score_df


df = zs_level_to_score()


def write_df_to_excel(df, sheet_name, start_col, end_col):
    '''
    将传入的DataFrame写入指定的excel的sheet中,from P to Y
    df:作业和实验为分数制的DataFrame
    '''
    wb = load_workbook(adress)
    ws = wb[sheet_name]
    logging.info('successfully open the {}'.format(adress))
    new_colums = ['z{}'.format(i) for i in range(1, 6)] + ['s{}'.format(i) for i
                                                           in
                                                           range(1, 6)]  # 新列名
    fill_in_new_column_name(ws, new_colums, start_col, end_col)
    score_nar = df.values  # 将 DataFrame 转为二维数组
    # 填充作业和实验的分数
    for row, col_values in enumerate(score_nar, start=2):
        for col_value, col in zip(col_values, range(start_col, end_col + 1)):
            ws.cell(row, col).value = col_value
    wb.save(adress)
    logging.info('successfully run the write_df_to_excel')


def fill_in_new_column_name(ws, new_colums, start_col, end_col):
    '''
    填入新列名
    '''
    for col, col_name in zip(range(start_col, end_col + 1), new_colums):
        ws.cell(1, col).value = col_name
    logging.info(
        'sucessfully fill in the col name: {}~{}'.format(start_col, end_col))


def fill_in_Zscore_Sscore(sheet_name):
    '''
    填入平时成绩（25%），实验成绩（25%）,from AA to AB
    '''
    wb = load_workbook(adress)
    ws = wb[sheet_name]
    new_colums = ['平时成绩(25%)', '实验成绩(25%)']
    fill_in_new_column_name(ws, new_colums, index('AA') + 1, index('AB') + 1)
    score_nar = zs_level_to_score().values
    # 填入成绩
    for row, col_values in enumerate(score_nar, start=2):
        ws.cell(row, index('AA') + 1).value = sum(col_values[:5])
        ws.cell(row, index('AB') + 1).value = sum(col_values[5:])
    wb.save(adress)
    logging.info('successfully fill in Z and S score')


def fill_in_test_score():
    '''
    从考试成绩表中计算考试成绩并填充，填入学生成绩表和考试成绩表中，并计算考试成绩(50%)和总成绩
    :return 返回缺考名单
    '''
    # 读取考试的题型细分成绩DataFrame
    test_score_df = sheet4.iloc[1:student_number + 1, index('C'):index('J') + 1]
    test_score_nar = test_score_df.values
    # 填充两个表的考试成绩
    wb = load_workbook(adress)
    ws = wb['考试成绩']
    ws.cell(2, index('K') + 1).value = '成绩'
    for row, col_values in enumerate(test_score_nar, start=3):
        col_values = col_values[1::2]
        ws.cell(row, index('K') + 1).value = sum(col_values)
    ws = wb['学生成绩']
    ws.cell(1, index('AE') + 1).value = '考试成绩'
    for row, col_values in enumerate(test_score_nar, start=2):
        col_values = col_values[1::2]
        ws.cell(row, index('AE') + 1).value = sum(col_values)
    wb.save(adress)
    # 填充学生成绩表的考试成绩(50%)
    ws.cell(1, index('AC') + 1).value = '考试成绩(50%)'
    for row, col_values in enumerate(test_score_nar, start=2):
        col_values = col_values[1::2]
        ws.cell(row, index('AC') + 1).value = sum(col_values) * 0.5
    wb.save(adress)
    # 填充学生成绩表的总成绩
    ws.cell(1, index('AD') + 1).value = '总成绩'
    for row in range(2, student_number + 2):  ##################
        sum_score = ws.cell(row, index('AA') + 1).value + ws.cell(row, index('AB') + 1).value + ws.cell(row, index('AC') + 1).value
        #由于浮点数存储不确定，使用该方法确保正确四舍五入
        if abs(sum_score - round(sum_score)) > 0.2:
            sum_score = round(sum_score + 0.2)
        else:
            sum_score = round(sum_score)
        print(sum_score)
        ws.cell(row, index('AD') + 1).value = sum_score
    wb.save(adress)
    # 统计缺考名单
    absent_row_list_ = [row for row, col_values in
                        enumerate(test_score_nar, start=2) if
                        all(np.isnan(x) for x in col_values[::2])]
    ws = wb['学生信息']
    absent_list = [ws.cell(row, 2).value for row in absent_row_list_]
    return absent_list
    logging.info('successfully run fill_in_test_score')


def fill_goal_in_test_score():
    '''
    填充三个目标的得分到考试成绩表中的M~O,并填充计算三个的平均值
    '''
    wb = load_workbook(adress, data_only=True)
    ws = wb['考试成绩']
    # 填充列名
    for i in range(1, 4):
        ws.cell(2, index('M') + i).value = '目标{}'.format(i)
    # 填充数据
    for row in range(3, student_number + 3):
        ws.cell(row, index('M') + 1).value = ((ws.cell(row, index(
            'E') + 1).value if ws.cell(row,
                                       index('E') + 1).value != None else 0) + (
                                                  ws.cell(row, index(
                                                      'G') + 1).value if ws.cell(
                                                      row, index(
                                                          'G') + 1).value != None else 0)) / 2
        ws.cell(row, index('N') + 1).value = (ws.cell(row, index(
            'I') + 1).value if ws.cell(row,
                                       index('I') + 1).value != None else 0) / 2
        ws.cell(row, index('O') + 1).value = (ws.cell(row, index(
            'C') + 1).value if ws.cell(row,
                                       index('C') + 1).value != None else 0) / 2
    # 计算填充平均值
    for i in range(1, 4):
        ws.cell(student_number + 3, index('M') + i).value = '平均分数'
    lsts = [[], [], []]
    for row in range(3, student_number + 3):
        for lst, i in zip(lsts, range(1, 4)):
            lst.append(ws.cell(row, index('M') + i).value)
    for col, ls in zip(range(index('M') + 1, index('O') + 2), lsts):
        ws.cell(student_number + 4, col).value = sum(ls) / len(ls)
    wb.save(adress)
    logging.info('successfully run fill_goal_in_test_score')


def fill_in_zm_and_sm():
    '''
    填充作业目标和实验目标得分
    '''
    wb = load_workbook(adress)
    ws = wb['学生成绩']
    for i in range(1, 3):
        ws.cell(1, index('AG') + i).value = '作业目标{}'.format(i)
    for i in range(1, 3):
        ws.cell(1, index('AI') + i).value = '实验目标{}'.format(i)
    for row, col_values in enumerate(df.values, start=2):
        list1 = col_values[:2]
        list2 = col_values[2:5]
        list3 = col_values[5:7]
        list4 = col_values[7:]
        ws.cell(row, index('AG') + 1).value = sum(list1)
        ws.cell(row, index('AG') + 2).value = sum(list2)
        ws.cell(row, index('AG') + 3).value = sum(list3)
        ws.cell(row, index('AG') + 4).value = sum(list4)
    wb.save(adress)
    logging.info('successfully run fill_in_zm_and_sm')


def summary_of_achievement_sheet_filled():
    '''填充达成度汇总表'''
    wb = load_workbook(adress)
    ws1 = wb['学生信息']
    ws2 = wb['大纲要求']
    ws3 = wb['学生成绩']
    ws4 = wb['考试成绩']
    ws5 = wb['达成度汇总']
    # 填充学号和姓名
    ws5.cell(2, 1).value = '学号'
    ws5.cell(2, 2).value = '姓名'
    for row in range(3, student_number + 3):
        ws5.cell(row, 1).value = ws1.cell(row - 1, 1).value
        ws5.cell(row, 2).value = ws1.cell(row - 1, 2).value
    # 填充平时作业
    ws5.merge_cells(start_row=1, start_column=3, end_row=1,
                    end_column=4)  # 合并单元格
    ws5.cell(1, 3).value = '平时作业'
    ws5.cell(2, 3).value = '目标1'
    ws5.cell(2, 4).value = '目标2'
    for row in range(3, student_number + 3):
        ws5.cell(row, 3).value = ws3.cell(row - 1, index('AG') + 1).value
        ws5.cell(row, 4).value = ws3.cell(row - 1, index('AH') + 1).value
    # 填充实验目标
    ws5.merge_cells(start_row=1, start_column=5, end_row=1,
                    end_column=6)  # 合并单元格
    ws5.cell(1, 5).value = '实验作业'
    ws5.cell(2, 5).value = '目标1'
    ws5.cell(2, 6).value = '目标2'
    for row in range(3, student_number + 3):
        ws5.cell(row, 5).value = ws3.cell(row - 1, index('AI') + 1).value
        ws5.cell(row, 6).value = ws3.cell(row - 1, index('AJ') + 1).value
    # 填充期末考试
    ws5.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    ws5.cell(1, 7).value = '期末考试'
    for i in range(1, 4):
        ws5.cell(2, 6 + i).value = '目标{}'.format(i)
    for row in range(3, student_number + 3):
        ws5.cell(row, index('G') + 1).value = ws4.cell(row,
                                                       index('M') + 1).value
        ws5.cell(row, index('H') + 1).value = ws4.cell(row,
                                                       index('N') + 1).value
        ws5.cell(row, index('I') + 1).value = ws4.cell(row,
                                                       index('O') + 1).value
    # 填充课程目标
    ws5.merge_cells(start_row=1, start_column=(index('J') + 1), end_row=1,
                    end_column=(index('L') + 1))
    ws5.cell(1, index('J') + 1).value = '课程目标'
    for i in range(1, 4):
        ws5.cell(2, index('J') + i).value = '目标{}'.format(i)
    for row in range(3, student_number + 3):
        ws5.cell(row, index('J') + 1).value = ws5.cell(row, index(
            'C') + 1).value + ws5.cell(row, index('E') + 1).value + ws5.cell(
            row, index('G') + 1).value
        ws5.cell(row, index('K') + 1).value = ws5.cell(row, index(
            'D') + 1).value + ws5.cell(row, index('F') + 1).value + ws5.cell(
            row, index('H') + 1).value
        ws5.cell(row, index('L') + 1).value = ws5.cell(row,
                                                       index('I') + 1).value
    # 填充汇总
    ws5.cell(2, index('M') + 1).value = '汇总'
    for row in range(3, student_number + 3):
        ws5.cell(row, index('M') + 1).value = ws5.cell(row, index(
            'J') + 1).value + ws5.cell(row, index('K') + 1).value + ws5.cell(
            row, index('L') + 1).value
    # 填充目标达成度
    ws5.merge_cells(start_row=1, start_column=(index('N') + 1), end_row=1,
                    end_column=(index('P') + 1))
    ws5.cell(1, index('N') + 1).value = '目标达成度'
    for i in range(1, 4):
        ws5.cell(2, index('N') + i).value = '目标{}'.format(i)
    for row in range(3, student_number + 3):
        ws5.cell(row, index('N') + 1).value = round(
            ws5.cell(row, index('J') + 1).value / ws2.cell(3, index(
                'F') + 1).value, 2)
        ws5.cell(row, index('O') + 1).value = round(
            ws5.cell(row, index('K') + 1).value / ws2.cell(4, index(
                'F') + 1).value, 2)
        ws5.cell(row, index('P') + 1).value = round(
            ws5.cell(row, index('L') + 1).value / ws2.cell(5, index(
                'F') + 1).value, 2)
    # 填充毕业达成要求
    ws5.merge_cells(start_row=1, start_column=(index('Q') + 1), end_row=1,
                    end_column=(index('S') + 1))
    ws5.cell(1, index('Q') + 1).value = '毕业达成度'
    ws5.cell(2, index('Q') + 1).value = '毕业要求1.2'
    ws5.cell(2, index('R') + 1).value = '毕业要求3.3'
    ws5.cell(2, index('S') + 1).value = '毕业要求5.1'
    for row in range(3, student_number + 3):
        ws5.cell(row, index('Q') + 1).value = ws5.cell(row,
                                                       index('N') + 1).value
        ws5.cell(row, index('R') + 1).value = ws5.cell(row,
                                                       index('O') + 1).value
        ws5.cell(row, index('S') + 1).value = ws5.cell(row,
                                                       index('P') + 1).value
    wb.save(adress)


def main():
    write_df_to_excel(df, '学生成绩', index('P') + 1, index('Y') + 1)
    fill_in_Zscore_Sscore('学生成绩')
    absent_list = fill_in_test_score()
    print('缺考名单为: ', end='')
    print(absent_list)
    fill_in_zm_and_sm()
    fill_goal_in_test_score()
    summary_of_achievement_sheet_filled()


if __name__ == '__main__':
    main()
