import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import pandas as pd

# 设置中文字体显示
plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei","SimSun"]

# 数据加载地址
adress = r"C:\Users\lenovo\Desktop\课程开发平台试验excel表.xlsx"

try:
    # 加载工作簿和工作表
    wb = load_workbook(adress)
    ws = wb['学生成绩']

    # 读取数据到DataFrame (假设表头在第一行)
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    # 获取列名(表头)
    headers = [cell.value for cell in ws[1]]

    # 创建DataFrame
    df = pd.DataFrame(data, columns=headers)

    # 打印示例数据
    print(f"数据加载成功，共{len(df)}条记录")
    print(f"示例数据: {df.head().to_dict('records')}")

    # 创建直方图并获取区间人数
    fig, ax = plt.subplots(figsize=(10, 6))
    n, bins, patches = ax.hist(df['总成绩'], bins=[0, 60, 70, 80, 90, 101],
                               color='skyblue', edgecolor='black', alpha=0.7)

    # 统计各区间人数(改为列表)
    counts = [int(count) for count in n]

    # 打印统计结果
    print("\n各分数段人数统计:")
    intervals = ["0-60分", "60-70分", "70-80分", "80-90分", "90-100分"]
    for interval, count in zip(intervals, counts):
        print(f"{interval}: {count}人")

    # 在直方图上添加数值标签
    for i in range(len(patches)):
        height = patches[i].get_height()
        ax.text(patches[i].get_x() + patches[i].get_width()/2., height + 0.5,
                f'{counts[i]}', ha='center', va='bottom')

    # 设置图表属性
    ax.set_title('学生总成绩分布直方图')
    ax.set_xlabel('分数区间')
    ax.set_ylabel('学生人数')
    ax.grid(axis='y', linestyle='--', alpha=0.7)

    # 显示图形
    plt.tight_layout()
    plt.show()

except FileNotFoundError:
    print(f"错误: 文件 '{adress}' 未找到")
except Exception as e:
    print(f"发生未知错误: {e}")